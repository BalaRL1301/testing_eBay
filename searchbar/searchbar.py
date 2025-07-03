from selenium import webdriver
from selenium.common import exceptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
import logging
import os

#logging
logging.basicConfig(
    filename = "log/searchbar.log",
    filemode = 'a',
    format = "%(asctime)s - %(levelname)s - %(message)s",
    level = logging.INFO
)

def ebay_searchbar_test(driver):
    logging.info("SEARCH TESTING INFORMATION")
    print("searching test initialized")
    assert "Electronics, Cars, Fashion, Collectibles & More | eBay" in driver.title, "Title does not matech !!"

    data = load_workbook('/Users/balarl/Desktop/testing1/eBay/searchbar/search_items.xlsx')
    sheet = data.active

    for row in sheet.iter_rows(min_row = 2, values_only = True):
        item = row[0]
        search_bar = driver.find_element(By.ID, 'gh-ac')
        search_bar.clear()
        search_bar.send_keys(item)
        search_bar.send_keys(Keys.RETURN)
        time.sleep(3)

        titles = driver.find_elements(By.CLASS_NAME, "s-item__title")
        found = False
        for title in titles:
            title_text = title.text.strip().lower()
            if item.lower() in title_text:
                found = True
                logging.info("listed", item)
                break
        # use of assertion
        try:
            assert found, f"Item '{item}' not listed in the search results"
            logging.info(f" Assertion Passed: Item '{item}' was found in the search results.")
        except AssertionError as ae:
            logging.error(f" Assertion Failed: {ae}")
            print(ae)

        driver.back()
    print("searchbar test is complelted")
    driver.quit()