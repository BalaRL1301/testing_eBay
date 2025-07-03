from selenium import webdriver
from selenium.common import exceptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
#from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
import time
import logging
import os
#next page to rediect
from searchbar.searchbar import ebay_searchbar_test

# Create directories for log and screenshots
os.makedirs("log", exist_ok=True)
os.makedirs("screenshots", exist_ok=True)

# Configure logging to file
logging.basicConfig(
    filename='log/login.log',
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# List to hold logs for HTML report
html_logs = []

def log_html(level, message, driver=None):
    """Log message and optionally include a screenshot"""
    timestamp = time.strftime('%Y-%m-%d %H-%M-%S')
    screenshot_path = ""
    if driver:
        screenshot_filename = f"screenshots/{timestamp}.png"
        try:
            driver.save_screenshot(screenshot_filename)
            screenshot_path = f'<a href="{screenshot_filename}" target="_blank">View Screenshot</a>'
        except Exception as e:
            screenshot_path = f"(screenshot failed: {e})"

    log_entry = f"<tr><td>{timestamp}</td><td>{level}</td><td>{message}<br>{screenshot_path}</td></tr>"
    html_logs.append(log_entry)
    getattr(logging, level.lower())(message)

def write_html_report():
    """Write all collected logs into a simple HTML file"""
    html_template = """
    <html>
    <head><title>eBay Login Automation Report</title></head>
    <body>
        <h2>eBay Login Automation Test Report</h2>
        <table border="1" cellpadding="5" cellspacing="0">
            <thead>
                <tr><th>Timestamp</th><th>Level</th><th>Message</th></tr>
            </thead>
            <tbody>
                {rows}
            </tbody>
        </table>
    </body>
    </html>
    """
    with open("login_report.html", "w", encoding="utf-8") as f:
        f.write(html_template.format(rows="\n".join(html_logs)))


#    -----  exel file creation  -----

output_file = 'login_output.xlsx'
if not os.path.exists(output_file):
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Login Result"
    ws_out.append(["Email", "Status"])
    wb_out.save(output_file)

def save_result(email, status, snippet=""):
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active
    ws_out.append([email, status])
    wb_out.save(output_file)

#   -----   Main test   -----
def ebay_login_test(excel_file='user_Credentials.xlsx'):
    print("Login test initialized")
    log_html("INFO", "LOGIN TEST STARTED")

    try:
        driver = webdriver.Chrome()
        driver.maximize_window()
    except exceptions.WebDriverException as e:
        log_html("ERROR", f"WebDriver initialization failed: {e}")
        return

    try:
        driver.get("https://www.ebay.com/")
    except Exception as e:
        log_html("ERROR", f"Failed to load eBay homepage: {e}", driver)
        driver.quit()
        return

    try:
        data = load_workbook(excel_file)
        sheet = data.active
    except FileNotFoundError:
        log_html("ERROR", f"Excel file '{excel_file}' not found.", driver)
        driver.quit()
        return
    except InvalidFileException as e:
        log_html("ERROR", f"Invalid Excel file: {e}", driver)
        driver.quit()
        return

    try:
        driver.find_element(By.CSS_SELECTOR, '#gh > nav > div.gh-nav__left-wrap > span.gh-identity > span > a').click()
    except exceptions.NoSuchElementException:
        log_html("ERROR", "Login button not found on eBay homepage.", driver)
        driver.quit()
        return

    for row in sheet.iter_rows(min_row= 2, max_row = 5, values_only=True):
        email, password = row

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'userid')))
            user_input = driver.find_element(By.ID, 'userid')
            user_input.clear()
            user_input.send_keys(email)
            driver.find_element(By.ID, 'signin-continue-btn').click()
            time.sleep(2)

            # CAPTCHA - manual

            if driver.find_elements(By.ID, 'signin-error-msg'):
                log_html("ERROR", f"Account not found - {email}", driver)
                save_result(email, "Account not found")
                try:
                    driver.execute_script(f"alert('Account not found for email: {email}');")
                    time.sleep(3)
                    driver.switch_to.alert.accept()
                except exceptions.NoAlertPresentException:
                    log_html("WARNING", "Expected alert not present.", driver)
                driver.get("https://signin.ebay.com/")
                continue

            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'pass')))
            driver.find_element(By.ID, 'pass').send_keys(password)
            driver.find_element(By.ID, 'sgnBt').click()
            log_html("INFO", f"Login attempted for email: {email}")
            save_result(email, "valid email")
            time.sleep(2)

            if driver.find_elements(By.ID, 'signin-error-msg'):
                log_html("ERROR", f"Incorrect password for email {email}", driver)
                save_result(email, "Incorrect password for this email")
                driver.back()
            else:
                log_html("INFO", f"Login successful - {email}")
                save_result(email, "Login successful")
            time.sleep(2)

        except (exceptions.NoSuchElementException, exceptions.TimeoutException) as e:
            log_html("ERROR", f"Element not found or timeout for {email}: {e}", driver)
        except exceptions.ElementNotInteractableException as e:
            log_html("ERROR", f"Element not interactable for {email}: {e}", driver)
        except exceptions.StaleElementReferenceException as e:
            log_html("ERROR", f"Stale element reference for {email}: {e}", driver)
        except Exception as e:
            log_html("ERROR", f"Unexpected exception for {email}: {e}", driver)

    log_html("INFO", "LOGIN TEST COMPLETED")
    print("Login test completed .. Check log/login.log, login_report.html and login_output.xlsx  ")


    write_html_report()
    ebay_searchbar_test(driver)
    #driver.quit()
